import io
import os
import base64
import qrcode
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from .models import (
    HappySheetResponse, PreAssessmentResponse, PostAssessmentResponse,
    BARSTechnicalResponse, BARSBehaviouralResponse, BARSLeadershipResponse,
)
from .competencies import (
    MT_COMPETENCIES, GET_COMPETENCIES,
    HAPPY_SHEET_QUESTIONS, SURVEY_META,
)

# ─────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────

def get_client_ip(request):
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded:
        return x_forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def get_competencies(role):
    return MT_COMPETENCIES if role == 'MT' else GET_COMPETENCIES


def make_qr_b64(url):
    qr = qrcode.QRCode(version=1, box_size=8, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#1a2d5a', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()


def get_base_url(request):
    base = getattr(settings, 'BASE_URL', '')
    if base and base != 'http://localhost:8000':
        return base.rstrip('/')
    scheme = 'https' if request.is_secure() else 'http'
    return f"{scheme}://{request.get_host()}"


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('admin_logged_in'):
            return redirect('/admin-login/')
        return view_func(request, *args, **kwargs)
    return wrapper


# ─────────────────────────────────────────
# Welcome / Index
# ─────────────────────────────────────────

def welcome(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        emp_code = request.POST.get('emp_code', '').strip()
        grade = request.POST.get('grade', '').strip()
        role = request.POST.get('role', '').strip()
        survey_type = request.POST.get('survey_type', '').strip()

        if not all([name, emp_code, role, survey_type]):
            return render(request, 'surveys/welcome.html', {
                'error': 'Please fill in all required fields.',
                'name': name, 'emp_code': emp_code, 'grade': grade,
                'role': role,
            })

        request.session['user_name'] = name
        request.session['user_emp_code'] = emp_code
        request.session['user_grade'] = grade
        request.session['user_role'] = role

        survey_map = {
            'happy-sheet': f'/survey/{role.lower()}/happy-sheet/',
            'pre-assessment': f'/survey/{role.lower()}/pre-assessment/',
            'post-assessment': f'/survey/{role.lower()}/post-assessment/',
            'bars-technical': f'/survey/{role.lower()}/bars-technical/',
            'bars-behavioural': f'/survey/{role.lower()}/bars-behavioural/',
            'bars-leadership': f'/survey/{role.lower()}/bars-leadership/',
        }
        url = survey_map.get(survey_type)
        if url:
            return redirect(url)
        return render(request, 'surveys/welcome.html', {'error': 'Invalid survey type selected.'})

    return render(request, 'surveys/welcome.html')


# ─────────────────────────────────────────
# Happy Sheet — MT & GET
# ─────────────────────────────────────────

def happy_sheet(request, role):
    role = role.upper()
    if role not in ('MT', 'GET'):
        return redirect('/')

    meta = SURVEY_META['happy-sheet']
    prefill = {
        'name': request.session.get('user_name', ''),
        'emp_code': request.session.get('user_emp_code', ''),
    }

    if request.method == 'POST':
        try:
            obj = HappySheetResponse(
                role=role,
                employee_id=request.POST.get('employee_id', '').strip(),
                name=request.POST.get('name', '').strip(),
                date=request.POST.get('date'),
                program_name=request.POST.get('program_name', '').strip(),
                vendor=request.POST.get('vendor', '').strip(),
                competency=request.POST.get('competency', '').strip(),
                mode=request.POST.get('mode', '').strip(),
                trainer=request.POST.get('trainer', '').strip(),
                q1=int(request.POST.get('q1', 0)),
                q2=int(request.POST.get('q2', 0)),
                q3=int(request.POST.get('q3', 0)),
                q4=int(request.POST.get('q4', 0)),
                q5=int(request.POST.get('q5', 0)),
                q6=int(request.POST.get('q6', 0)),
                q7=int(request.POST.get('q7', 0)),
                q8=int(request.POST.get('q8', 0)),
                q9=int(request.POST.get('q9', 0)),
                q10=int(request.POST.get('q10', 0)),
                nps_score=int(request.POST.get('nps_score', 0)),
                open_feedback=request.POST.get('open_feedback', '').strip(),
                supervisor=request.POST.get('supervisor', '').strip(),
                hr_bp=request.POST.get('hr_bp', '').strip(),
                ip_address=get_client_ip(request),
            )
            obj.save()
            return redirect(f'/survey/{role.lower()}/success/?type=happy-sheet')
        except Exception as e:
            return render(request, 'surveys/happy_sheet.html', {
                'role': role, 'meta': meta, 'questions': HAPPY_SHEET_QUESTIONS,
                'error': f'Submission failed: {e}', 'prefill': prefill,
            })

    return render(request, 'surveys/happy_sheet.html', {
        'role': role,
        'meta': meta,
        'questions': HAPPY_SHEET_QUESTIONS,
        'prefill': prefill,
    })


# ─────────────────────────────────────────
# Pre-Assessment
# ─────────────────────────────────────────

def pre_assessment(request, role):
    role = role.upper()
    if role not in ('MT', 'GET'):
        return redirect('/')

    comps = get_competencies(role)['pre_post']
    meta = SURVEY_META['pre-assessment']
    prefill = {
        'name': request.session.get('user_name', ''),
        'emp_code': request.session.get('user_emp_code', ''),
        'grade': request.session.get('user_grade', ''),
    }

    if request.method == 'POST':
        try:
            ratings = {}
            for c in comps:
                rating_val = request.POST.get(f'rating_{c["code"]}')
                evidence = request.POST.get(f'evidence_{c["code"]}', '').strip()
                if rating_val:
                    ratings[c['code']] = {
                        'name': c['name'],
                        'domain': c['domain'],
                        'rating': int(rating_val),
                        'evidence': evidence,
                    }

            obj = PreAssessmentResponse(
                role=role,
                rater_type=request.POST.get('rater_type', 'candidate'),
                employee_id=request.POST.get('employee_id', '').strip(),
                name=request.POST.get('name', '').strip(),
                role_grade=request.POST.get('role_grade', '').strip(),
                dept=request.POST.get('dept', '').strip(),
                manager_name=request.POST.get('manager_name', '').strip(),
                date=request.POST.get('date'),
                competency_ratings=ratings,
                ip_address=get_client_ip(request),
            )
            obj.save()
            return redirect(f'/survey/{role.lower()}/success/?type=pre-assessment')
        except Exception as e:
            return render(request, 'surveys/assessment.html', {
                'role': role, 'meta': meta, 'competencies': comps,
                'survey_type': 'pre', 'error': str(e), 'prefill': prefill,
            })

    return render(request, 'surveys/assessment.html', {
        'role': role,
        'meta': meta,
        'competencies': comps,
        'survey_type': 'pre',
        'prefill': prefill,
    })


# ─────────────────────────────────────────
# Post-Assessment
# ─────────────────────────────────────────

def post_assessment(request, role):
    role = role.upper()
    if role not in ('MT', 'GET'):
        return redirect('/')

    comps = get_competencies(role)['pre_post']
    meta = SURVEY_META['post-assessment']
    prefill = {
        'name': request.session.get('user_name', ''),
        'emp_code': request.session.get('user_emp_code', ''),
        'grade': request.session.get('user_grade', ''),
    }

    if request.method == 'POST':
        try:
            ratings = {}
            for c in comps:
                rating_val = request.POST.get(f'rating_{c["code"]}')
                evidence = request.POST.get(f'evidence_{c["code"]}', '').strip()
                if rating_val:
                    ratings[c['code']] = {
                        'name': c['name'],
                        'domain': c['domain'],
                        'rating': int(rating_val),
                        'evidence': evidence,
                    }

            obj = PostAssessmentResponse(
                role=role,
                rater_type=request.POST.get('rater_type', 'candidate'),
                employee_id=request.POST.get('employee_id', '').strip(),
                name=request.POST.get('name', '').strip(),
                training_completed=request.POST.get('training_completed', '').strip(),
                date=request.POST.get('date'),
                manager_name=request.POST.get('manager_name', '').strip(),
                days_since_training=int(request.POST.get('days_since_training') or 0) or None,
                competency_ratings=ratings,
                ip_address=get_client_ip(request),
            )
            obj.save()
            return redirect(f'/survey/{role.lower()}/success/?type=post-assessment')
        except Exception as e:
            return render(request, 'surveys/assessment.html', {
                'role': role, 'meta': meta, 'competencies': comps,
                'survey_type': 'post', 'error': str(e), 'prefill': prefill,
            })

    return render(request, 'surveys/assessment.html', {
        'role': role,
        'meta': meta,
        'competencies': comps,
        'survey_type': 'post',
        'prefill': prefill,
    })


# ─────────────────────────────────────────
# BARS helper
# ─────────────────────────────────────────

def _handle_bars(request, role, bars_type, ModelClass):
    role = role.upper()
    if role not in ('MT', 'GET'):
        return redirect('/')

    comp_key = f'bars_{bars_type}'
    comps = get_competencies(role)[comp_key]
    meta = SURVEY_META[f'bars-{bars_type}']
    prefill = {'name': request.session.get('user_name', ''), 'emp_code': request.session.get('user_emp_code', '')}

    if request.method == 'POST':
        try:
            ratings = {}
            for c in comps:
                rating_val = request.POST.get(f'rating_{c["code"]}')
                evidence = request.POST.get(f'evidence_{c["code"]}', '').strip()
                if rating_val:
                    ratings[c['code']] = {
                        'name': c['name'],
                        'rating': int(rating_val),
                        'evidence': evidence,
                    }

            obj = ModelClass(
                role=role,
                employee_id=request.POST.get('employee_id', '').strip(),
                name=request.POST.get('name', '').strip(),
                obs_number=request.POST.get('obs_number', '30'),
                period=request.POST.get('period', '').strip(),
                supervisor=request.POST.get('supervisor', '').strip(),
                date=request.POST.get('date'),
                ratings=ratings,
                ip_address=get_client_ip(request),
            )
            obj.save()
            return redirect(f'/survey/{role.lower()}/success/?type=bars-{bars_type}')
        except Exception as e:
            return render(request, 'surveys/bars.html', {
                'role': role, 'meta': meta, 'competencies': comps,
                'bars_type': bars_type, 'error': str(e), 'prefill': prefill,
            })

    return render(request, 'surveys/bars.html', {
        'role': role, 'meta': meta, 'competencies': comps,
        'bars_type': bars_type, 'prefill': prefill,
    })


def bars_technical(request, role):
    return _handle_bars(request, role, 'technical', BARSTechnicalResponse)


def bars_behavioural(request, role):
    return _handle_bars(request, role, 'behavioural', BARSBehaviouralResponse)


def bars_leadership(request, role):
    return _handle_bars(request, role, 'leadership', BARSLeadershipResponse)


# ─────────────────────────────────────────
# Success Page
# ─────────────────────────────────────────

def success(request, role):
    survey_type = request.GET.get('type', 'survey')
    meta = SURVEY_META.get(survey_type, {'title': 'Survey', 'icon': '✅', 'color': '#28a745', 'subtitle': ''})
    return render(request, 'surveys/success.html', {
        'role': role.upper(),
        'meta': meta,
        'survey_type': survey_type,
    })


# ─────────────────────────────────────────
# QR Code Generation
# ─────────────────────────────────────────

def qr_codes(request):
    if not request.session.get('admin_logged_in'):
        return redirect('/admin-login/')

    base = get_base_url(request)
    survey_links = []

    for role_key, role_label in [('mt', 'MT (L11 | Band 4)'), ('get', 'GET (L08T | Band 5)')]:
        for stype, smeta in SURVEY_META.items():
            url = f"{base}/survey/{role_key}/{stype}/"
            qr_b64 = make_qr_b64(url)
            survey_links.append({
                'role': role_label,
                'role_key': role_key.upper(),
                'type': stype,
                'title': smeta['title'],
                'subtitle': smeta['subtitle'],
                'icon': smeta['icon'],
                'color': smeta['color'],
                'url': url,
                'qr_b64': qr_b64,
            })

    return render(request, 'surveys/qr_codes.html', {'survey_links': survey_links, 'base_url': base})


# ─────────────────────────────────────────
# Admin Login / Logout
# ─────────────────────────────────────────

def admin_login(request):
    if request.session.get('admin_logged_in'):
        return redirect('/admin-dashboard/')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        if (username == settings.ADMIN_USERNAME and
                password == settings.ADMIN_PASSWORD):
            request.session['admin_logged_in'] = True
            request.session['admin_username'] = username
            return redirect('/admin-dashboard/')
        error = 'Invalid credentials. Please try again.'

    return render(request, 'surveys/admin_login.html', {'error': error})


def admin_logout(request):
    request.session.flush()
    return redirect('/admin-login/')


# ─────────────────────────────────────────
# Admin Dashboard
# ─────────────────────────────────────────

@admin_required
def admin_dashboard(request):
    stats = {
        'happy_sheet': HappySheetResponse.objects.count(),
        'pre_assessment': PreAssessmentResponse.objects.count(),
        'post_assessment': PostAssessmentResponse.objects.count(),
        'bars_technical': BARSTechnicalResponse.objects.count(),
        'bars_behavioural': BARSBehaviouralResponse.objects.count(),
        'bars_leadership': BARSLeadershipResponse.objects.count(),
    }
    stats['total'] = sum(stats.values())

    recent_happy = HappySheetResponse.objects.order_by('-submitted_at')[:5]
    recent_pre = PreAssessmentResponse.objects.order_by('-submitted_at')[:5]
    recent_bars_t = BARSTechnicalResponse.objects.order_by('-submitted_at')[:5]

    # Role breakdown
    mt_happy = HappySheetResponse.objects.filter(role='MT').count()
    get_happy = HappySheetResponse.objects.filter(role='GET').count()

    base = get_base_url(request)

    return render(request, 'surveys/admin_dashboard.html', {
        'stats': stats,
        'recent_happy': recent_happy,
        'recent_pre': recent_pre,
        'recent_bars_t': recent_bars_t,
        'mt_happy': mt_happy,
        'get_happy': get_happy,
        'base_url': base,
    })


# ─────────────────────────────────────────
# Admin – View Responses
# ─────────────────────────────────────────

@admin_required
def admin_responses(request, survey_type):
    role_filter = request.GET.get('role', '')
    valid_types = {
        'happy-sheet': (HappySheetResponse, 'Happy Sheet Responses'),
        'pre-assessment': (PreAssessmentResponse, 'Pre-Assessment Responses'),
        'post-assessment': (PostAssessmentResponse, 'Post-Assessment Responses'),
        'bars-technical': (BARSTechnicalResponse, 'BARS Technical Responses'),
        'bars-behavioural': (BARSBehaviouralResponse, 'BARS Behavioural Responses'),
        'bars-leadership': (BARSLeadershipResponse, 'BARS Leadership Responses'),
    }

    if survey_type not in valid_types:
        return redirect('/admin-dashboard/')

    Model, title = valid_types[survey_type]
    qs = Model.objects.all()
    if role_filter in ('MT', 'GET'):
        qs = qs.filter(role=role_filter)

    return render(request, 'surveys/admin_responses.html', {
        'responses': qs,
        'title': title,
        'survey_type': survey_type,
        'role_filter': role_filter,
        'meta': SURVEY_META.get(survey_type, {}),
    })


# ─────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────

@admin_required
def export_excel(request, survey_type):
    role_filter = request.GET.get('role', '')

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A2D5A')
    sub_fill = PatternFill('solid', fgColor='2E4B8F')
    sub_font = Font(bold=True, color='FFFFFF', size=10)
    alt_fill = PatternFill('solid', fgColor='E8F0FE')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

    def style_row(ws, row, cols, alt=False):
        fill = alt_fill if alt else None
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            if fill:
                cell.fill = fill
            cell.alignment = left
            cell.border = thin_border

    if survey_type == 'happy-sheet':
        qs = HappySheetResponse.objects.all()
        if role_filter in ('MT', 'GET'):
            qs = qs.filter(role=role_filter)

        ws = wb.active
        ws.title = 'Happy Sheet Responses'
        headers = ['#', 'Role', 'Emp ID', 'Name', 'Date', 'Program', 'Vendor',
                   'Competency', 'Mode', 'Trainer', 'Q1', 'Q2', 'Q3', 'Q4', 'Q5',
                   'Q6', 'Q7', 'Q8', 'Q9', 'Q10', 'Avg Score', 'NPS',
                   'Open Feedback', 'Supervisor', 'HR-BP', 'Submitted At']
        ws.append(headers)
        style_header(ws, 1, len(headers))

        for i, r in enumerate(qs, 1):
            row_data = [
                i, r.role, r.employee_id, r.name, str(r.date), r.program_name,
                r.vendor, r.competency, r.mode, r.trainer,
                r.q1, r.q2, r.q3, r.q4, r.q5, r.q6, r.q7, r.q8, r.q9, r.q10,
                r.average_score(), r.nps_score, r.open_feedback, r.supervisor,
                r.hr_bp, r.submitted_at.strftime('%Y-%m-%d %H:%M')
            ]
            ws.append(row_data)
            style_row(ws, i + 1, len(headers), alt=(i % 2 == 0))

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['X'].width = 40
        for col in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
            ws.column_dimensions[col].width = 7

    elif survey_type in ('pre-assessment', 'post-assessment'):
        Model = PreAssessmentResponse if survey_type == 'pre-assessment' else PostAssessmentResponse
        qs = Model.objects.all()
        if role_filter in ('MT', 'GET'):
            qs = qs.filter(role=role_filter)

        ws = wb.active
        ws.title = f"{'Pre' if survey_type == 'pre-assessment' else 'Post'}-Assessment"

        sample = qs.first()
        comp_codes = list(sample.competency_ratings.keys()) if sample else []

        base_headers = ['#', 'Role', 'Rater', 'Emp ID', 'Name', 'Grade/Dept', 'Manager', 'Date']
        comp_headers = [f"{c}\n(Rating)" for c in comp_codes] + [f"{c}\n(Evidence)" for c in comp_codes]
        headers = base_headers + comp_headers + ['Submitted At']
        ws.append(headers)
        style_header(ws, 1, len(headers))

        for i, r in enumerate(qs, 1):
            base_row = [i, r.role, r.rater_type, r.employee_id, r.name,
                        getattr(r, 'role_grade', '') or getattr(r, 'training_completed', ''),
                        r.manager_name, str(r.date)]
            ratings_row = [r.competency_ratings.get(c, {}).get('rating', '') for c in comp_codes]
            evidence_row = [r.competency_ratings.get(c, {}).get('evidence', '') for c in comp_codes]
            ws.append(base_row + ratings_row + evidence_row + [r.submitted_at.strftime('%Y-%m-%d %H:%M')])
            style_row(ws, i + 1, len(headers), alt=(i % 2 == 0))

        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['G'].width = 25

    elif survey_type in ('bars-technical', 'bars-behavioural', 'bars-leadership'):
        model_map = {
            'bars-technical': BARSTechnicalResponse,
            'bars-behavioural': BARSBehaviouralResponse,
            'bars-leadership': BARSLeadershipResponse,
        }
        Model = model_map[survey_type]
        qs = Model.objects.all()
        if role_filter in ('MT', 'GET'):
            qs = qs.filter(role=role_filter)

        ws = wb.active
        ws.title = f"BARS {survey_type.split('-')[1].title()}"

        sample = qs.first()
        comp_codes = list(sample.ratings.keys()) if sample else []

        base_headers = ['#', 'Role', 'Emp ID', 'Name', 'Obs', 'Period', 'Supervisor', 'Date']
        comp_headers = ([f"{c}\n(Rating)" for c in comp_codes] +
                        [f"{c}\n(Evidence)" for c in comp_codes] + ['Avg Rating'])
        headers = base_headers + comp_headers + ['Submitted At']
        ws.append(headers)
        style_header(ws, 1, len(headers))

        for i, r in enumerate(qs, 1):
            base_row = [i, r.role, r.employee_id, r.name, f"{r.obs_number} days",
                        r.period, r.supervisor, str(r.date)]
            ratings_row = [r.ratings.get(c, {}).get('rating', '') for c in comp_codes]
            evidence_row = [r.ratings.get(c, {}).get('evidence', '') for c in comp_codes]
            ws.append(base_row + ratings_row + evidence_row + [r.average_rating(), r.submitted_at.strftime('%Y-%m-%d %H:%M')])
            style_row(ws, i + 1, len(headers), alt=(i % 2 == 0))

        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['G'].width = 25

    # Freeze top row on all sheets
    for sheet in wb.worksheets:
        sheet.freeze_panes = 'A2'

    # Title row
    title_ws = wb.create_sheet('Info', 0)
    title_ws['A1'] = 'JSW Motors Training Evaluation Data Export'
    title_ws['A1'].font = Font(bold=True, size=14, color='1A2D5A')
    title_ws['A2'] = f'Survey Type: {survey_type.replace("-", " ").title()}'
    title_ws['A3'] = f'Role Filter: {role_filter or "All Roles"}'
    title_ws['A4'] = f'Exported At: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    title_ws.column_dimensions['A'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"JSW_Training_{survey_type}_{role_filter or 'ALL'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
